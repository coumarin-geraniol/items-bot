from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime
from database.database import get_user_info
from openpyxl.styles import PatternFill, Border, Side, Font


async def create_excel_file(user_id, cargo_id, cart_items, total_price):
    unique_time_str = datetime.now().strftime("%Y%m%d-%H%M%S")

    user_info = get_user_info(user_id)
    if not cart_items:
        return None  # Возвращаем None, если корзина пуста

    # Создаем новую книгу Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Cart"

    # Добавляем информацию о пользователе в первые строки файла
    ws.append(['User ID', user_info['id']])
    ws.append(['TG ID', user_info['tg_id']])
    ws.append(['TG Username', '@'+user_info['tg_username']])
    ws.append(['Username', user_info['username']])
    ws.append([])  # Пустая строка для разделения

    # Заполняем заголовки столбцов
    headers = ['Order ID', 'Item Name', 'Quantity', 'Price per Item', 'Total']
    ws.append(headers)
    column_widths = {'A': 15, 'B': 20, 'C': 15, 'D': 20, 'E': 20}
    for column, width in column_widths.items():
        ws.column_dimensions[column].width = width


    # Добавляем данные по каждому товару
    for item in cart_items:
        ws.append([item['order_id'], item['name'], item['quantity'], item['price_per_item'], item['item_total_price']])

    # Добавляем общую стоимость заказа в конец файла
    ws.append([])
    ws.append(['', '', '', 'Total Price:', total_price])

    # Выравниваем ячейки по центру
    for row in ws.iter_rows(min_row=1, max_col=5, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

    # Сохраняем файл
    filename = f"export/docs/cart_{cargo_id}-{unique_time_str}.xlsx"
    wb.save(filename)
    return filename
